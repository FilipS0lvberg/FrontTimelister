from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Color, Alignment, Border, Side, PatternFill
import io

def create_prosjektregnskap1(year: int, juridisknavn: str, project_number: int, project_title: str, godkjent_periode: str, status_dato: str, status: str, godkjent_dato: str, kategori: str, selskap_vansker: str, cost_options):
    wb = Workbook()

    ###################### CREATE FRONT PAGE (REPORT) (Sheet 0) ##########################
    ws0 = wb.active
    ws0.sheet_view.showGridLines = False
    ws0.title = "Rapport"

    # Light gray
    fill_lg = PatternFill(start_color='00C0C0C0', 
            end_color='00C0C0C0', 
            fill_type='solid')

    ws0["B2"] = f"Prosjektregnskap {year}"
    ws0["B2"].font = Font(size=20, bold=True)

    ws0.column_dimensions["B"].width = 50
    ws0.column_dimensions["C"].width = 50


    ws0["B4"] = "Juridisk selskapsnavn"
    ws0["B5"] = "Prosjektnummer"
    ws0["B6"] = "Prosjekttittel (Beskrivelse)"

    ws0["B8"] = "Godkjenningsperiode (dato fra - dato til)"
    ws0["B9"] = "Prosjektstatusdato"
    ws0["B10"] = "Prosjektstatus"
    ws0["B11"] = "Søknad godkjent dato"
    ws0["B12"] = "Kategori"
    ws0["B13"] = "Var selskapet i vansker ved godkjenningstidspunkt målt etter siste avlagte årsregnskap?"
    ws0["B13"].alignment = Alignment(wrap_text=True)
    
    ### C kolonne
    
    ws0["C4"] = juridisknavn
    ws0["C4"].alignment = Alignment(horizontal="right")
    ws0["C5"] = project_number
    ws0["C5"].alignment = Alignment(horizontal="right")
    ws0["C6"] = project_title
    ws0["C6"].alignment = Alignment(horizontal="right")
    ws0["C8"] = godkjent_periode
    ws0["C8"].alignment = Alignment(horizontal="right")
    ws0["C9"] = status_dato
    ws0["C9"].alignment = Alignment(horizontal="right")
    ws0["C10"] = status
    ws0["C10"].alignment = Alignment(horizontal="right")
    ws0["C11"] = godkjent_dato
    ws0["C11"].alignment = Alignment(horizontal="right")
    ws0["C12"] = kategori
    ws0["C12"].alignment = Alignment(horizontal="right")
    ws0["C13"] = selskap_vansker
    ws0["C13"].alignment = Alignment(horizontal="right")

    # Adding borders
    start_cell = "B4"
    end_cell = "C6"
    cell_range = ws0[start_cell:end_cell]

    border_A = Border(bottom=Side(style="thin"), top=Side(style="thin"), left=Side(style="thin"), right=Side(style="thin"))
    for row in cell_range:
        for cell in row:
            cell.border = border_A

    start_cell = "B8"
    end_cell = "C13"
    cell_range = ws0[start_cell:end_cell]

    for row in cell_range:
        for cell in row:
            cell.border = border_A


    # Begynnelse av Rapport sammendrag
    ws0["B16"] = "Prosjektkostnader"
    ws0["B16"].font = Font(size=12, bold=True)

    border_u = Border(bottom=Side(style="thin"))
    ws0["B16"].border = border_u
    ws0["c16"].border = border_u

    ## cost_opts = ["Timekostnader", "Prosjektkostnader", "Bruk av eget utstyr", "Kapitalkostnader"]
    counter = 0
    for option in cost_options:
        row = 17 + counter
        if option == "Timekostnader":
            ws0[f"B{row}"] = "Timer"
            ws0[f"C{row}"] = "=Timekostnader!F25"
            counter += 1
        
        if option == "Prosjektkostnader":
            ws0[f"B{row}"] = "Prosjektkostnader"
            ws0[f"C{row}"] = "=Prosjektkostnader!F31"
            counter += 1

        if option == "Bruk av eget utstyr":
            ws0[f"B{row}"] = "Bruk av eget utstyr"
            ws0[f"C{row}"] = "='Bruk av eget utstyr'!H16"
            counter += 1

        if option == "Kapitalkostnader":
            ws0[f"B{row}"] = "Kapitalkostnader"
            ws0[f"C{row}"] = "='Kapitalkostnader'!H16"
            counter += 1


    ws0["B20"].border = border_u
    ws0["C20"].border = border_u

    ws0["B21"] = "Sum prosjektkostnader"
    ws0["B21"].font = Font(size=12, bold=True)
    ws0["C21"] = "=SUM(C17:C20)"
    ws0["C21"].font = Font(size=12, bold=True)

    ws0["B23"].border = border_u
    ws0["C23"].border = border_u
    ws0["B24"] = "Grunnlag for støtte - overføres til skattebergningen"
    ws0["B24"].font = Font(size=12, bold=True)
    ws0["C24"] = "=MIN(C21,25000000)"
    ws0["C24"].font = Font(size=12, bold=True)
    ws0["B25"] = "Støttesats"
    ws0["C25"] = "19 %"
    ws0["C25"].alignment = Alignment(horizontal="right")

    ws0["B26"].border = border_u
    ws0["C26"].border = border_u
    ws0["B27"] = "Skattefradrag"
    ws0["B27"].font = Font(size=12, bold=True)
    ws0["C27"] = "=C24*0.19"
    ws0["C27"].font = Font(size=12, bold=True)
    ws0["B27"].border = border_u
    ws0["C27"].border = border_u



    ###########################################################################################################
    #### Time kostnader ark
    if "Timekostnader" in cost_options:
        ws1 = wb.create_sheet("Timekostnader")
        ws1.sheet_view.showGridLines = False

        ws1["B2"] = f"Timekostnader"
        ws1["B2"].font = Font(size=20, bold=True)

        ws1.column_dimensions['B'].width = 30
        ws1.column_dimensions['C'].width = 20
        ws1.column_dimensions['D'].width = 15
        ws1.column_dimensions['E'].width = 15
        ws1.column_dimensions['F'].width = 15

        ws1["B4"] = "Selskap"
        ws1["F4"] = juridisknavn
        ws1["F4"].alignment = Alignment(horizontal="right")

        ws1["B5"] = "Prosjektnummer"
        ws1["F5"] = project_number
        ws1["F5"].alignment = Alignment(horizontal="right")

        ws1["B6"] = "Prosjekttittel"
        ws1["F6"] = project_title
        ws1["F6"].alignment = Alignment(horizontal="right")

        start_cell = "B4"
        end_cell = "F6"
        cell_range = ws1[start_cell:end_cell]

        for row in cell_range:
            for cell in row:
                cell.border = border_u


        # Timer table
        ws1["B9"] = "Navn"
        ws1["B9"].font = Font(size=12, bold=True)
        ws1["B9"].fill = fill_lg

        ws1["C9"] = "Brutto årslønn"
        ws1["C9"].font = Font(size=12, bold=True)
        ws1["C9"].fill = fill_lg

        ws1["D9"] = "Timepris"
        ws1["D9"].font = Font(size=12, bold=True)
        ws1["D9"].fill = fill_lg

        ws1["E9"] = "Timer"
        ws1["E9"].font = Font(size=12, bold=True)
        ws1["E9"].fill = fill_lg

        ws1["F9"] = "Timekostnad"
        ws1["F9"].font = Font(size=12, bold=True)
        ws1["F9"].fill = fill_lg

        for i in range(15):
            ws1[f"D{10+i}"] = f"=MIN(C{10+i}*0.0012,700)"
            ws1[f"D{10+i}"].alignment = Alignment(horizontal="right")
            ws1[f"F{10+i}"] = f"=E{10+i}*D{10+i}"
            ws1[f"F{10+i}"].alignment = Alignment(horizontal="right")
            ws1[f"F{10+i}"].number_format = '#,##0.00kr' 


        ws1["B25"] = "Sum"
        ws1["B25"].fill = fill_lg
        ws1["C25"].fill = fill_lg
        ws1["D25"].fill = fill_lg
        ws1["E25"] = "=SUM(E10:E24)"
        ws1["E25"].fill = fill_lg
        ws1["F25"] = "=SUM(F10:F24)"
        ws1["F25"].number_format = '#,##0.00kr'
        ws1["F25"].fill = fill_lg

        start_cell = "B9"
        end_cell = "F25"
        cell_range = ws1[start_cell:end_cell]

        for row in cell_range:
            for cell in row:
                cell.border = border_A

        ws1["B28"] = "*Brutto årslønn er kun spesifisert dersom den er under 584 000 kr/år"
        ws1["B29"] = "** Anvendt årslønn er oppgitt etter siste lønnskorrigering før 31/12 inneværende skatteår, ref FSFIN § 16-40-6 tredje ledd."




    ###############################################################################################################################
    # page for prosjektkostnader
    if "Prosjektkostnader" in cost_options:
        ws2 = wb.create_sheet("Prosjektkostnader")
        ws2.sheet_view.showGridLines = False

        ws2["B2"] = f"Timekostnader"
        ws2["B2"].font = Font(size=20, bold=True)

        ws2.column_dimensions['B'].width = 30
        ws2.column_dimensions['C'].width = 20
        ws2.column_dimensions['D'].width = 20
        ws2.column_dimensions['E'].width = 20
        ws2.column_dimensions['F'].width = 20
        ws2.column_dimensions['G'].width = 20

        ws2["B4"] = "Selskap"
        ws2["E4"] = juridisknavn
        ws2["E4"].alignment = Alignment(horizontal="right")

        ws2["B5"] = "Prosjektnummer"
        ws2["E5"] = project_number
        ws2["E5"].alignment = Alignment(horizontal="right")

        ws2["B6"] = "Prosjekttittel"
        ws2["E6"] = project_title
        ws2["E6"].alignment = Alignment(horizontal="right")

        start_cell = "B4"
        end_cell = "E6"
        cell_range = ws2[start_cell:end_cell]

        for row in cell_range:
            for cell in row:
                cell.border = border_u

        # Timer table
        ws2["B9"] = "Leverandør"
        ws2["B9"].font = Font(size=12, bold=True)
        ws2["B9"].fill = fill_lg

        ws2["C9"] = "Kommentar"
        ws2["C9"].font = Font(size=12, bold=True)
        ws2["C9"].fill = fill_lg

        ws2["D9"] = "Fakturanr/ID"
        ws2["D9"].font = Font(size=12, bold=True)
        ws2["D9"].fill = fill_lg

        ws2["E9"] = "Kontonummer"
        ws2["E9"].font = Font(size=12, bold=True)
        ws2["E9"].fill = fill_lg

        ws2["F9"] = "Beløp, eks. MVA"
        ws2["F9"].font = Font(size=12, bold=True)
        ws2["F9"].fill = fill_lg

        ws2["G9"] = "Forfallsdato"
        ws2["G9"].font = Font(size=12, bold=True)
        ws2["G9"].fill = fill_lg

        for i in range(20):
            ws2[f"F{10+i}"].alignment = Alignment(horizontal="right")
            ws2[f"F{10+i}"].number_format = '#,##0.00kr' 

        ws2["B31"] = "Sum"
        ws2["B31"].fill = fill_lg
        ws2["C31"].fill = fill_lg
        ws2["D31"].fill = fill_lg
        ws2["E31"].fill = fill_lg
        ws2["F31"] = "=SUM(F10:F30)"
        ws2["F31"].number_format = '#,##0.00kr'
        ws2["F31"].fill = fill_lg
        ws2["G31"].fill = fill_lg

        start_cell = "B9"
        end_cell = "G31"
        cell_range = ws2[start_cell:end_cell]

        for row in cell_range:
            for cell in row:
                cell.border = border_A


    ####################################################################################################################
    ## Page for bruk av eget utstyr
    if "Bruk av eget utstyr" in cost_options:
        ws3 = wb.create_sheet("Bruk av eget utstyr")
        ws3.sheet_view.showGridLines = False

        ws3["B2"] = f"Timekostnader"
        ws3["B2"].font = Font(size=20, bold=True)

        ws3.column_dimensions['A'].width = 5
        ws3.column_dimensions['B'].width = 20
        ws3.column_dimensions['C'].width = 20
        ws3.column_dimensions['D'].width = 20
        ws3.column_dimensions['E'].width = 20
        ws3.column_dimensions['F'].width = 20
        ws3.column_dimensions['G'].width = 20
        ws3.column_dimensions['H'].width = 20
        ws3.column_dimensions['I'].width = 20

        ws3["B4"] = "Selskap"
        ws3["E4"] = juridisknavn
        ws3["E4"].alignment = Alignment(horizontal="right")

        ws3["B5"] = "Prosjektnummer"
        ws3["E5"] = project_number
        ws3["E5"].alignment = Alignment(horizontal="right")

        ws3["B6"] = "Prosjekttittel"
        ws3["E6"] = project_title
        ws3["E6"].alignment = Alignment(horizontal="right")

        start_cell = "B4"
        end_cell = "E6"
        cell_range = ws3[start_cell:end_cell]

        for row in cell_range:
            for cell in row:
                cell.border = border_u

        # Timer table
        ws3["B9"] = "Kostnadstype (type renter)"
        ws3["B9"].font = Font(size=12, bold=True)
        ws3["B9"].fill = fill_lg

        ws3["C9"] = "Kreditor"
        ws3["C9"].font = Font(size=12, bold=True)
        ws3["C9"].fill = fill_lg

        ws3["D9"] = "Bilagsnummer"
        ws3["D9"].font = Font(size=12, bold=True)
        ws3["D9"].fill = fill_lg

        ws3["E9"] = "Gjeld relatert til prosjektet"
        ws3["E9"].font = Font(size=12, bold=True)
        ws3["E9"].fill = fill_lg

        ws3["F9"] = "Rentesats"
        ws3["F9"].font = Font(size=12, bold=True)
        ws3["F9"].fill = fill_lg

        ws3["G9"] = "Dato"
        ws3["G9"].font = Font(size=12, bold=True)
        ws3["G9"].fill = fill_lg

        ws3["H9"] = "Beløp"
        ws3["H9"].font = Font(size=12, bold=True)
        ws3["H9"].fill = fill_lg

        ws3["I9"] = "Kommentar"
        ws3["I9"].font = Font(size=12, bold=True)
        ws3["I9"].fill = fill_lg

        for i in range(5):
            ws3[f"F{10+i}"].alignment = Alignment(horizontal="right")
            ws3[f"F{10+i}"].number_format = '#,##0.00kr' 

        ws3["B16"] = "Sum"
        ws3["B16"].fill = fill_lg
        ws3["C16"].fill = fill_lg
        ws3["D16"].fill = fill_lg
        ws3["E16"].fill = fill_lg
        ws3["F16"].fill = fill_lg
        ws3["G16"].fill = fill_lg
        ws3["H16"].fill = fill_lg
        ws3["H16"] = "=SUM(H10:H15)"
        ws3["H16"].number_format = '#,##0.00kr'
        ws3["I16"].fill = fill_lg

        start_cell = "B9"
        end_cell = "I16"
        cell_range = ws3[start_cell:end_cell]

        for row in cell_range:
            for cell in row:
                cell.border = border_A

        ws3["B17"] = "Belastning på prosjekt: Hvor mange dager/ timer etc. er utstyret benyttet på prosjektet"
        ws3["B18"] = "Annen belastning: Hvor mange dager/timer etc. er utstyret benyttet til andre formål enn prosjektet. Herunder også dødtid. Obs: om utstyret utelukkende er benyttet til prosjektet kan hele avskrivingskosten medtas i prosjektregnskapet. "


    ####################################################################################################################
    ## Page for Kapitalkostnader
    if "Kapitalkostnader" in cost_options:
        ws4 = wb.create_sheet("Kapitalkostnader")
        ws4.sheet_view.showGridLines = False

        ws4["B2"] = f"Timekostnader"
        ws4["B2"].font = Font(size=20, bold=True)

        ws4.column_dimensions['A'].width = 3
        ws4.column_dimensions['B'].width = 20
        ws4.column_dimensions['C'].width = 25
        ws4.column_dimensions['D'].width = 25
        ws4.column_dimensions['E'].width = 30
        ws4.column_dimensions['F'].width = 30
        ws4.column_dimensions['G'].width = 20
        ws4.column_dimensions['H'].width = 15
        ws4.column_dimensions['I'].width = 15

        ws4["B4"] = "Selskap"
        ws4["E4"] = juridisknavn
        ws4["E4"].alignment = Alignment(horizontal="right")

        ws4["B5"] = "Prosjektnummer"
        ws4["E5"] = project_number
        ws4["E5"].alignment = Alignment(horizontal="right")

        ws4["B6"] = "Prosjekttittel"
        ws4["E6"] = project_title
        ws4["E6"].alignment = Alignment(horizontal="right")

        start_cell = "B4"
        end_cell = "E6"
        cell_range = ws4[start_cell:end_cell]

        for row in cell_range:
            for cell in row:
                cell.border = border_u

        # Timer table
        ws4["B9"] = "Utstyr"
        ws4["B9"].font = Font(size=12, bold=True)
        ws4["B9"].fill = fill_lg

        ws4["C9"] = "Enhet (timer/dager etc.)"
        ws4["C9"].font = Font(size=12, bold=True)
        ws4["C9"].fill = fill_lg

        ws4["D9"] = "Belastning på prosjekt"
        ws4["D9"].font = Font(size=12, bold=True)
        ws4["D9"].fill = fill_lg

        ws4["E9"] = "Annen belastning pluss dødtid"
        ws4["E9"].font = Font(size=12, bold=True)
        ws4["E9"].fill = fill_lg

        ws4["F9"] = "Årets skattemessige avskrivning"
        ws4["F9"].font = Font(size=12, bold=True)
        ws4["F9"].fill = fill_lg

        ws4["G9"] = "Eiendelsnummer"
        ws4["G9"].font = Font(size=12, bold=True)
        ws4["G9"].fill = fill_lg

        ws4["H9"] = "Sum"
        ws4["H9"].font = Font(size=12, bold=True)
        ws4["H9"].fill = fill_lg

        ws4["I9"] = "Kommentar"
        ws4["I9"].font = Font(size=12, bold=True)
        ws4["I9"].fill = fill_lg

        for i in range(5):
            ws4[f"F{10+i}"].alignment = Alignment(horizontal="right")
            ws4[f"F{10+i}"].number_format = '#,##0.00kr' 

        ws4["B16"] = "Sum"
        ws4["B16"].fill = fill_lg
        ws4["C16"].fill = fill_lg
        ws4["D16"].fill = fill_lg
        ws4["E16"].fill = fill_lg
        ws4["F16"].fill = fill_lg
        ws4["G16"].fill = fill_lg
        ws4["H16"].fill = fill_lg
        ws4["H16"] = "=SUM(H10:H15)"
        ws4["H16"].number_format = '#,##0.00kr'
        ws4["I16"].fill = fill_lg

        start_cell = "B9"
        end_cell = "I16"
        cell_range = ws4[start_cell:end_cell]

        for row in cell_range:
            for cell in row:
                cell.border = border_A

        ws4["B17"] = "*Kapitalkostnader kan i følge Skatte ABC 2020 s. 566 inngå som del av fradragsgrunnlaget gitt at de ikke relateres til lønnskostnader og forutsatt at kapitalkostnadene har vært nødvendige for gjennomførselen av prosjektet. "
        #ws4["B17"].alignment = Alignment(wrap_text=True)



    # save workbook
    # wb.save("testbook.xlsx")
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()



if __name__ == "__main__":
    cost_opts = ["Timekostnader", "Prosjektkostnader", "Bruk av eget utstyr", "Kapitalkostnader"]
    create_prosjektregnskap1(2025, "Front Innovation AS", 1283, "Nytt test prosjekt", godkjent_periode="01.01.2025 - 01.01.2026", status_dato="01.02.2025", status="Godkjent", godkjent_dato="06.09.2024", kategori="Utviklingsprosjekt eksperimentell utvikling", selskap_vansker="Nei", cost_options=cost_opts)