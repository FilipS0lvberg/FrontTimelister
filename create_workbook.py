from openpyxl.workbook import *
from openpyxl.styles import Font, Color, Alignment, Border, Side, PatternFill
from supporting_functions import get_dates_in_quarter, get_first_letters_dict
import io


def create_workbook1(year: int, quarter: int, selskap: str, prosjekttittel: str, prosjektnummer: str, prosjektleder: str, workers: list, prosject_description: str, num_workpackages: int):
    # Create workbook
    wb = Workbook()

    ###################### CREATE FRONT PAGE (REPORT) (Sheet 0) ##########################
    ws0 = wb.active
    ws0.sheet_view.showGridLines = False
    ws0.title = "Rapport"

    ws0["B2"] = f"Timelister {year} - Q{quarter}"
    ws0["B2"].font = Font(size=20, bold=True)

    ws0["B3"] = "Selskap"
    ws0["B4"] = "Prosjekttittel"
    ws0["B5"] = "Prosjektnummer"
    ws0["B6"] = "Prosjektleder"

    start_cell = "B3"
    end_cell = "I6"
    cell_range = ws0[start_cell:end_cell]

    border = Border(bottom=Side(style="thin"))
    for row in cell_range:
        for cell in row:
            cell.border = border

    ws0.merge_cells("D3:I3")
    ws0.merge_cells("D4:I4")
    ws0.merge_cells("D5:I5")
    ws0.merge_cells("D6:I6")

    ws0["D3"].alignment = Alignment(horizontal="left")
    ws0["D4"].alignment = Alignment(horizontal="left")
    ws0["D5"].alignment = Alignment(horizontal="left")
    ws0["D6"].alignment = Alignment(horizontal="left")

    ws0["D3"] = selskap
    ws0["D4"] = prosjekttittel
    ws0["D5"] = prosjektnummer
    ws0["D6"] = prosjektleder

    quarter_map = {1: [1, 2, 3], 2: [4, 5, 6], 3: [7, 8, 9], 4: [10, 11, 12]}

    ws0["C9"] = "Måned"
    ws0["B10"] = "Navn"
    ws0["C10"] = quarter_map[quarter][0]
    ws0["D10"] = quarter_map[quarter][1]
    ws0["E10"] = quarter_map[quarter][2]
    ws0["F10"] = "Sum totalt"
    ws0["F10"].fill = PatternFill(
            start_color='FFE2EFDA',
            end_color='FFE2EFDA',
            fill_type='solid'
        )


    ################# CREATING PROJECT DESCRIPTION PAGE (SHEET 1) ############################
    ws1 = wb.create_sheet("Prosjektbeskrivelse")
    ws1.sheet_view.showGridLines = False

    # Light gray
    fill = PatternFill(start_color='00C0C0C0', 
                end_color='00C0C0C0', 
                fill_type='solid')

    ws1["B2"] = f"Prosjektbeskrivelse"
    ws1["B2"].font = Font(size=18, bold=True)

    ws1.column_dimensions['C'].width = 80

    # 5. Write a large amount of descriptive text into a cell
    ws1["C4"] = prosject_description

    # 6. Ensure the text wraps instead of overflowing
    ws1["C4"].alignment = Alignment(wrap_text=True)


    ########################### ADDING WORKERS TIME LIST PAGES ###############################
    for index, worker in enumerate(workers):
        new_sheet = wb.create_sheet(worker)
        new_sheet.sheet_view.showGridLines = False
        new_sheet.column_dimensions["A"].width = 5
        new_sheet.column_dimensions["B"].width = 12
        new_sheet.column_dimensions["C"].width = 12
        new_sheet.column_dimensions["D"].width = 12
        new_sheet.column_dimensions["E"].width = 12

        new_sheet.column_dimensions["F"].width = 20
        new_sheet.column_dimensions["G"].width = 20
        new_sheet.column_dimensions["H"].width = 20
        new_sheet.column_dimensions["I"].width = 20
        new_sheet.column_dimensions["J"].width = 20
        new_sheet.column_dimensions["K"].width = 20
        new_sheet.column_dimensions["L"].width = 20
        new_sheet.column_dimensions["M"].width = 20
        new_sheet.column_dimensions["N"].width = 20
        new_sheet.column_dimensions["O"].width = 20

        new_sheet["B2"] = f"Navn"
        new_sheet["B2"].font = Font(size=12, bold=True)
        new_sheet["B3"] = "Selskap"
        new_sheet["B4"] = "Prosjekt"
        new_sheet["B5"] = "Prosjektnummer"
        new_sheet["B6"] = "Prosjektleder"

        start_cell = "B2"
        end_cell = "F6"
        cell_range = new_sheet[start_cell:end_cell]

        border1 = Border(bottom=Side(style="thin"))
        for row in cell_range:
            for cell in row:
                cell.border = border1

        new_sheet.merge_cells("D2:F2")
        new_sheet.merge_cells("D3:F3")
        new_sheet.merge_cells("D4:F4")
        new_sheet.merge_cells("D5:F5")
        new_sheet.merge_cells("D6:F6")

        new_sheet["D3"].alignment = Alignment(horizontal="left")
        new_sheet["D4"].alignment = Alignment(horizontal="left")
        new_sheet["D5"].alignment = Alignment(horizontal="left")
        new_sheet["D6"].alignment = Alignment(horizontal="left")

        new_sheet["D2"] = "" #f"{worker}"
        new_sheet["D2"].fill = fill
        new_sheet["D3"] = selskap
        new_sheet["D4"] = prosjekttittel
        new_sheet["D5"] = prosjektnummer
        new_sheet["D6"] = prosjektleder

        new_sheet.merge_cells("D8:F8")
        new_sheet["D8"] = "Arbeidspakker"
        new_sheet["D8"].font = Font(size=20, bold=True)

        new_sheet["B9"] = "Navn"
        new_sheet["B9"].font = Font(size=12, bold=True)
        new_sheet["B9"].border = border1
        new_sheet["C9"] = "Dato"
        new_sheet["C9"].font = Font(size=12, bold=True)
        new_sheet["C9"].border = border1
        new_sheet["D9"] = "Måned"
        new_sheet["D9"].font = Font(size=12, bold=True)
        new_sheet["D9"].border = border1
        new_sheet["E9"] = "Sum dag"
        new_sheet["E9"].font = Font(size=12, bold=True)
        new_sheet["E9"].border = border1


        for pa in range(num_workpackages):
            new_sheet.cell(row=9, column=6+pa, value=f"Arbeidspakke {pa+1}").font = Font(size=12, bold=True)
            new_sheet.cell(row=9, column=6+pa).border = border1


        ### HERE STARTS THE PROPER TABLE SETUP ###
        border2 = Border(left=Side(style="thin"))
        dates_in_quarter = get_dates_in_quarter(year, quarter)
        letter_map = get_first_letters_dict()

        # Let's say you want to start placing formulas from row 10 through row 30
        start_row1 = 10
        end_row1   = len(dates_in_quarter)

        # Navn kolonne
        for row in range(start_row1, end_row1 + 10):
            # Construct the formula so E10 becomes E{row}, E11 becomes E{row}, etc.
            # Example: =HVIS(E10<>"";$D$2;" ")
            formula = f'=IF(E{row}<>"",$D$2," ")'
            new_sheet.cell(row=row, column=2).value = formula

        # Dato kolonne
        for row_index, date in enumerate(dates_in_quarter):
            new_sheet.cell(row=row_index + start_row1, column=3).value = date[0]

        # Måned kolonne
        for row_index, date in enumerate(dates_in_quarter):
            new_sheet.cell(row=row_index + start_row1, column=4).value = int(date[0].split(".")[1])
            new_sheet.cell(row=row_index + start_row1, column=4).alignment = Alignment(horizontal="center")

        # Dag navn kolonne
        new_sheet.cell(row=9, column=5 + num_workpackages + 1).value = "Ukedag"
        new_sheet.cell(row=9, column=5 + num_workpackages + 1).alignment = Alignment(horizontal="center")
        new_sheet.cell(row=9, column=5 + num_workpackages + 1).font = Font(size=12, bold=True)
        new_sheet.cell(row=9, column=5 + num_workpackages + 1).border = Border(bottom=Side(style="thin"))
        for row_index, date in enumerate(dates_in_quarter):
            columnday = 5 + num_workpackages + 1
            new_sheet.cell(row=row_index + start_row1, column=columnday).value = date[2]
            new_sheet.cell(row=row_index + start_row1, column=columnday).alignment = Alignment(horizontal="center")

        # Sum dag kolonne
        for row in range(start_row1, end_row1 + 10):
            end_col = letter_map[5 + num_workpackages]
            formula = f'=IF(SUM(F{row}:{end_col}{row})>0,SUM(F{row}:{end_col}{row}),"")'
            new_sheet.cell(row=row, column=5).value = formula
            new_sheet.cell(row=row, column=5).alignment = Alignment(horizontal="center")

        # Time infyll per arbeidspakke kolonner
        for column1 in range(6, 6 + num_workpackages):
            for row1 in range(start_row1, end_row1 + 10):
                new_sheet.cell(row=row1, column=column1).border = border2
                new_sheet.cell(row=row1, column=column1).alignment = Alignment(horizontal="center")
                new_sheet.cell(row=row1, column=column1).fill = fill

        # light green
        fill2 = PatternFill(
            start_color='FFE2EFDA',
            end_color='FFE2EFDA',
            fill_type='solid'
        )
        
        for column in range(2, 2 + 4 + num_workpackages):
            for row in range(start_row1, end_row1 + 10):
                green_flag = dates_in_quarter[row - 10][1]

                if green_flag == 1:
                    new_sheet.cell(row=row, column=column).fill = fill2

        # Sum av timer
        end_col1 = letter_map[5 + num_workpackages]
        end_row1 = 10 + len(dates_in_quarter)
        new_sheet.cell(column=2 + 4 + 1, row=10 + len(dates_in_quarter)).value = "Sum timer totalt:"
        new_sheet.cell(column=2 + 4 + 1, row=10 + len(dates_in_quarter)).font = Font(size=12, bold=True)
        new_sheet.cell(column=2 + 4 + 1 + 1, row=10 + len(dates_in_quarter)).value = f"=SUM(F10:{end_col}{end_row1 - 1})"

        # Signatur linje
        border3 = Border(top=Side(style="thin"))
        new_sheet.merge_cells(f"C{end_row1 + 5}:D{end_row1 + 5}")
        new_sheet.merge_cells(f"G{end_row1 + 5}:H{end_row1 + 5}")
        new_sheet.merge_cells(f"C{end_row1 + 6}:D{end_row1 + 6}")
        new_sheet.merge_cells(f"G{end_row1 + 6}:H{end_row1 + 6}")

        new_sheet[f"C{end_row1 + 5}"].alignment = Alignment(horizontal="center")
        new_sheet[f"G{end_row1 + 5}"].alignment = Alignment(horizontal="center")
        new_sheet[f"C{end_row1 + 6}"].alignment = Alignment(horizontal="center")
        new_sheet[f"G{end_row1 + 6}"].alignment = Alignment(horizontal="center")

        new_sheet[f"C{end_row1 + 5}"].border = border3
        new_sheet[f"D{end_row1 + 5}"].border = border3
        new_sheet[f"H{end_row1 + 5}"].border = border3
        new_sheet[f"G{end_row1 + 5}"].border = border3

        new_sheet[f"C{end_row1 + 5}"] = "=D2"
        new_sheet[f"C{end_row1 + 6}"] = "Prosjektdeltaker"
        new_sheet[f"G{end_row1 + 5}"] = "=D6"
        new_sheet[f"G{end_row1 + 5}"] = "Prosjektleder"



    ######################## TILBAKE TIL REPORTING PAGE FOR Å SAMLE OG SAMMENSTILLE ALL DATAEN ##########################
    # Rapporteringsark ansattes timer per måned, per ansatt
    for index, worker in enumerate(workers):
        ws0[f"B{10+index+1}"] = f"={worker}!D2"
        ws0[f"C{10+index+1}"] = f"=SUMIF({worker}!D10:D{9+len(dates_in_quarter)},Rapport!C10,{worker}!E10:E{9+len(dates_in_quarter)})"
        ws0[f"D{10+index+1}"] = f"=SUMIF({worker}!D10:D{9+len(dates_in_quarter)},Rapport!D10,{worker}!E10:E{9+len(dates_in_quarter)})"
        ws0[f"E{10+index+1}"] = f"=SUMIF({worker}!D10:D{9+len(dates_in_quarter)},Rapport!E10,{worker}!E10:E{9+len(dates_in_quarter)})"
        ws0[f"F{10+index+1}"] = f"=SUM(C{10+index+1}:E{10+index+1})"
        ws0[f"F{10+index+1}"].fill = fill2

    # rapporteringsark, totale sum av alle ansattes timer
    ws0[f"B{10 + len(workers) + 1}"] = f"Sum Totalt"
    ws0[f"B{10 + len(workers) + 1}"].fill = fill2

    ws0[f"C{10 + len(workers) + 1}"] = f"=SUM(C{11}:C{10 + len(workers)})"
    ws0[f"C{10 + len(workers) + 1}"].fill = fill2

    ws0[f"D{10 + len(workers) + 1}"] = f"=SUM(D{11}:D{10 + len(workers)})"
    ws0[f"D{10 + len(workers) + 1}"].fill = fill2
    
    ws0[f"E{10 + len(workers) + 1}"] = f"=SUM(E{11}:E{10 + len(workers)})"
    ws0[f"E{10 + len(workers) + 1}"].fill = fill2

    ws0[f"F{10 + len(workers) + 1}"] = f"=SUM(F{11}:F{10 + len(workers)})"
    ws0[f"F{10 + len(workers) + 1}"].fill = fill2

    start_cell1 = "B10"
    end_cell1 = f"F{10 + len(workers) + 1}"
    cell_range2 = ws0[start_cell1:end_cell1]

    border4 = Border(bottom=Side(style="thin"), top=Side(style="thin"), left=Side(style="thin"), right=Side(style="thin"))
    for row in cell_range2:
        for cell in row:
            cell.border = border4


    # save workbook
    # wb.save("testbook.xlsx")
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()



if __name__ == "__main__":
    desc = (
        """Hovedmål: 
Prosjektets hovedmål er å kombinere brukervennligheten til low-code med dybden til et fullverdig programmeringsspråk for å tilby utviklere et intuitivt og brukervennlig verktøy, som skal være tidsbesparende sammenliknet med tradisjonell programvareutvikling. Løsningen skal:
- Tillate en person uten programmeringsbakgrunn å bruke det, med mindre enn 12 timer opplæring.
- Føre til 20% tidsbesparelse under utvikling av web API-er og sky-tjenester, sammenliknet med normal programvareutvikling.

Arbeidspakke 1: Utvikle grunnfunksjonalitet og grensesnitt
Aktivitet 1: Kartlegging og behovs analyse
Aktivitet 2: Kode grunnfunksjonene
Aktivitet 3: Utvikle grensesnittet
Aktivitet 4: Kombinere grunnfunksjoner med grensesnitt

Arbeidspakke 2: Utvikle digital infrastruktur
Aktivitet 1: Kartlegge løsninger for infrastruktur
Aktivitet 2: Utvikle server/ sky struktur

Arbeidspakke 3: Testing og iterasjon
Aktivitet 1: Privat beta testing
Aktivitet 2: Iterere"""
    )
    create_workbook1(2025, 3, "Front Innovation", "Test Prosjekt", "69420", "Filip S.H.", ["Filip", "Gashi"], desc, 4)